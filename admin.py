import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk
from PIL import Image, ImageTk
import bcrypt
import mysql.connector

def main_app(username, role):  # Add username parameter
    ctk.set_appearance_mode("system")
    ctk.set_default_color_theme("blue")
    
    app = ctk.CTk()
    app.geometry('930x700')
    app.title("Hệ thông quản lí thư viện") 
    app.resizable(False, False)
    
    # Tạo thanh menu và thêm nút đăng xuất
    def logout():
        app.destroy()
        import dangnhap
        dangnhap.main_app()
    
    menubar = tk.Menu(app)
    filemenu = tk.Menu(menubar, tearoff=0)  
    filemenu.add_command(label="Đăng xuất", command= logout)
    menubar.add_cascade(label="Tùy chọn", menu=filemenu)
    menubar.add_command(label="Xin chào, " + username, state="disabled")  # Add username to
    app.config(menu=menubar)
    
    # Tải và đặt hình nền
    img1 = ImageTk.PhotoImage(Image.open(r"D:\Study\HK I 2024\QTDL\Demo\assets\pattern.png"))
    l1 = ctk.CTkLabel(app, image=img1)
    l1.pack(fill="both", expand=True)
    
    # Tải ảnh banner 
    imgbanner = ctk.CTkImage(Image.open(r"D:\Study\HK I 2024\QTDL\Demo\assets\managerbanner.png"), size=(930, 125))
    l2 = ctk.CTkLabel(app, image=imgbanner, text="")
    l2.place(x=0, y=0)
    
    # Create a canvas and a scrollbar for content_frame
    canvas = tk.Canvas(app, width=1050, height=680, bg="#242424")
    canvas.place(x=50,y=165)
    
    # Scrollbars for the canvas
    scroll_y = tk.Scrollbar(app, orient="vertical", command=canvas.yview)
    scroll_y.place(x=1100, y=165, height=685)
    
    # Configure canvas scrolling
    canvas.configure(yscrollcommand=scroll_y.set)
    
    # Create the content frame
    content_frame = ctk.CTkFrame(canvas, width=1200, height=800, corner_radius=15, bg_color="#242424", fg_color="#242424")
    canvas.create_window((0, 0), window=content_frame, anchor="nw")
    
    def on_frame_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
    
    content_frame.bind("<Configure>", on_frame_configure)
    
    # Functions to update content frame
    def switch_to_author_management():
        #hàm tìm kiếm tác giả bằng mã hoặc tên tác giả
        def search_author( search_entry, combobox):
            # Lấy dữ liệu từ MySQL và hiển thị lên bảng
            # Kết nối đến cơ sở dữ liệu
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if combobox.get() == "Mã Tác Giả":
                mycursor.execute("SELECT * FROM TacGia WHERE MaTG LIKE %s", ('%'+search_entry.get()+'%',))
            elif combobox.get() == "Tên Tác Giả":
                mycursor.execute("SELECT * FROM TacGia WHERE TenTG LIKE %s", ('%'+search_entry.get()+'%',))
            elif combobox.get() == "Website":
                mycursor.execute("SELECT * FROM TacGia WHERE Website LIKE %s", ('%'+search_entry.get()+'%',))
            elif combobox.get() == "Ghi Chú":
                mycursor.execute("SELECT * FROM TacGia WHERE GhiChu LIKE %s", ('%'+search_entry.get()+'%',))
            myresult = mycursor.fetchall()
            # xóa dữ liệu cũ trong bảng
            for i in table.get_children():
                table.delete(i)
            # thêm dữ liệu mới vào bảng
            for row in myresult:
                table.insert("", "end", values=row)
            # đóng database
            mydb.close()
        
        #hàm thêm tác giả
        def add_author():
            pass
        label = ctk.CTkLabel(content_frame, text="Quản lí tác giả", font=("Century Gothic", 20), text_color="white")
        label.pack(pady=20)
        label.place(x=340, y=10)
        # Tạo ô tìm kiếm tác với thuộc tính tìm kiếm được ghi dịnh bằng combobox
        
        tg = ["Mã Tác Giả", "Tên Tác Giả", "Website", "Ghi Chú"]
        combobox = ctk.CTkComboBox(content_frame, values=tg, bg_color="#242424", border_color="#bcd5e8", corner_radius=10, button_color="#bcd5e8", button_hover_color="#60797d", dropdown_fg_color="#bcd5e8", dropdown_hover_color="#60797d", width=150, state="readonly")
        combobox.place(x=100, y=60)
        combobox.set("Tìm kiếm bằng")
        search_entry = ctk.CTkEntry(content_frame, width=300, placeholder_text="")
        search_entry.pack(pady=20)
        search_entry.place(x=260, y=60)
    
        # Tạo nút tìm kiếm
        search_button = ctk.CTkButton(content_frame, text="Tìm kiếm", width = 50 ,command=lambda: search_author(search_entry, combobox))
        search_button.pack(pady=20)
        search_button.place(x=570, y=60)
        
        # Hàm lấy tất cả tác giả
  
  
        def show_all_author():
            # Lấy dữ liệu từ MySQL và hiển thị lên bảng
            # Kết nối đến cơ sở dữ liệu
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            mycursor.execute("SELECT * FROM TacGia")
            myresult = mycursor.fetchall()
            # xóa dữ liệu cũ trong bảng
            for i in table.get_children():
                table.delete(i)
            # thêm dữ liệu mới vào bảng
            for row in myresult:
                table.insert("", "end", values=row)
            # đóng database
            mydb.close()
        
        #Tạo nút hiển thị tất cả tác giả
        show_all_button = ctk.CTkButton(content_frame, text="Hiển thị tất cả", width = 50, command=show_all_author)
        show_all_button.pack(pady=20)
        show_all_button.place(x=650, y=60)
        
        # Tạo bảng hiển thị tác giả, giới hạn số dòng hiển thị là 5 và tạo thêm thanh scroll để hiển thị tiếp các dòng conf lại
        columns = ("Mã Tác Giả", "Tên Tác Giả", "Website", "Ghi Chú")
        table = ttk.Treeview(content_frame, columns=columns, show="headings", height=11)
        table.place(x=117, y=120)
        for col in columns:
            table.heading(col, text = col)
            
        #thu nhỏ cột maTG
        table.column("Mã Tác Giả", width=100)
        table.column("Ghi Chú", width=300)
        # Cho chữ vô giữa ô hiển thị
        for col in columns:
            table.column(col, anchor="center")
            
        #Tạo thanh scroll cho bảng
        scroll = ttk.Scrollbar(content_frame, orient="vertical", command=table.yview)
        scroll.place(x=920, y=120, height=247)
        table.configure(yscrollcommand=scroll.set)
        # Lấy dữ liệu từ MySQL và hiển thị lên bảng
        show_all_author()
        
        #Tạo frame hiển thị thông tin tác giả
        info_frame = ctk.CTkFrame(content_frame, width=350, height=300, corner_radius=15, bg_color="#242424", fg_color="#242424")
        info_frame.place(x=95, y=300)
        
        #Tạo label và ô hiển thị thông tin tương ứng của tác giả
        label = ctk.CTkLabel(info_frame, text="Thông tin tác giả", font=("Century Gothic", 20), text_color="#ffffff")
        label.pack(pady=20)
        label.place(x=95, y=10)
        
        #Tạo label và ô hiển thị mã tác giả
        label = ctk.CTkLabel(info_frame, text="Mã tác giả", font=("Century Gothic", 15), text_color="#ffffff")
        label.pack(pady=10)
        label.place(x=10, y=40)
        maTG_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        maTG_entry.pack(pady=10)
        maTG_entry.place(x=10, y=70)
        
        #Tạo label và ô hiển thị tên tác giả
        label = ctk.CTkLabel(info_frame, text="Tên tác giả", font=("Century Gothic", 15), text_color="#ffffff")
        label.pack(pady=10)
        label.place(x=10, y=100)
        tenTG_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        tenTG_entry.pack(pady=10)
        tenTG_entry.place(x=10, y=130)
        
        #Tạo label và ô hiển thị website
        label = ctk.CTkLabel(info_frame, text="Website", font=("Century Gothic", 15), text_color="#ffffff")
        label.pack(pady=10)
        label.place(x=10, y=160)
        website_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        website_entry.pack(pady=10)
        website_entry.place(x=10, y=190)
        
        #Tạo label và ô hiển thị ghi chú
        label = ctk.CTkLabel(info_frame, text="Ghi chú", font=("Century Gothic", 15), text_color="#ffffff")
        label.pack(pady=10)
        label.place(x=10, y=220)
        ghiChu_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        ghiChu_entry.pack(pady=10)
        ghiChu_entry.place(x=10, y=250)
        
        # Bắt sự kiện khi click vào một dòng trong bảng sẽ hiển thị thông tin tương ứng của tác giả
        def on_click(event):
            item = table.selection()[0]
            maTG_entry.delete(0, "end")
            tenTG_entry.delete(0, "end")
            website_entry.delete(0, "end")
            ghiChu_entry.delete(0, "end")
            maTG_entry.insert(0, table.item(item, "values")[0])
            tenTG_entry.insert(0, table.item(item, "values")[1])
            website_entry.insert(0, table.item(item, "values")[2])
            ghiChu_entry.insert(0, table.item(item, "values")[3])
            
        table.bind("<ButtonRelease-1>", on_click)
        
        #Tạo frame chứa các nút thêm, cập nhật, xóa tác giả, reset thông tin tác giả
        button_frame = ctk.CTkFrame(content_frame, width=300, height=120, corner_radius=15, bg_color="#242424", fg_color="#242424")
        button_frame.place(x=500, y=300)
        
        #Thêm nút thêm và chức năng thêm tác giả
        def add_author():
            # Kết nối đến cơ sở dữ liệu
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            #Kiểm tra thông tin nhập vào đã đủ chưa hay đã tồn tại chưa, nếu không hiển thị message box cảnh báo
            if not maTG_entry.get() or not tenTG_entry.get() or not website_entry.get() or not ghiChu_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập đủ thông tin!")
            else:
                try:
                    mycursor.execute("INSERT INTO TacGia (MaTG, TenTG, Website, GhiChu) VALUES (%s, %s, %s, %s)", (maTG_entry.get(), tenTG_entry.get(), website_entry.get(), ghiChu_entry.get()))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Thêm tác giả thành công!")
                    show_all_author()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã tác giả đã tồn tại!")
            # Đóng kết nối database
            mydb.close()
        add_button = ctk.CTkButton(button_frame, text="Thêm", width = 100, command=add_author)
        add_button.pack(pady=30)
        add_button.place(x=0, y=25)        
                
        
        #Thêm nút cập nhật và chức nâng cập nhật thông tin tác giả
        def update_author():
            # Kết nối đến cơ sở dữ liệu
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            #Kiểm tra thông tin nhập vào đã đủ chưa hay đã tồn tại chưa, nếu không hiển thị message box cảnh báo
            if not maTG_entry.get() or not tenTG_entry.get() or not website_entry.get() or not ghiChu_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập đầy đủ thông tin tác giả cần cập nhật!")
            else:
                try:
                    mycursor.execute("UPDATE TacGia SET TenTG = %s, Website = %s, GhiChu = %s WHERE MaTG = %s", (tenTG_entry.get(), website_entry.get(), ghiChu_entry.get(), maTG_entry.get()))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Cập nhật thông tin tác giả thành công!")
                    show_all_author()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã tác giả không tồn tại!")
            # Đóng kết nối database
            mydb.close()
        update_button = ctk.CTkButton(button_frame, text="Cập nhật", width = 100, command=update_author)
        update_button.pack(pady=30)
        update_button.place(x=140, y=25)
        
        #Thêm nút xóa và chức năng xóa thông tin tác giả
        def delete_author():
            # Kết nối đến cơ sở dữ liệu
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            #Kiểm tra thông tin nhập vào đã đủ chưa hay đã tồn tại chưa, nếu không hiển thị message box cảnh báo
            if not maTG_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập mã tác giả cần xóa!")
            else:
                try:
                    mycursor.execute("DELETE FROM TacGia WHERE MaTG = %s", (maTG_entry.get(),))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Xóa tác giả thành công!")
                    show_all_author()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã tác giả không tồn tại!")
            # Đóng kết nối database
            mydb.close()
            reset_info()
        delete_button = ctk.CTkButton(button_frame, text="Xóa", width = 100, command=delete_author)
        delete_button.pack(pady=30)
        delete_button.place(x=140, y=80)
        
        #Thêm nút reset và chức năng reset thông tin tác giả
        def reset_info():
            maTG_entry.delete(0, "end")
            tenTG_entry.delete(0, "end")
            website_entry.delete(0, "end")
            ghiChu_entry.delete(0, "end")
        reset_button = ctk.CTkButton(button_frame, text="Reset", width = 100, command=reset_info)
        reset_button.pack(pady=30)
        reset_button.place(x=0, y=80)
        
        #Tạo frame chứa các nút xuất báo cáo thông kê
        report_frame = ctk.CTkFrame(content_frame, width=300, height=150, corner_radius=15, bg_color="#242424", fg_color="#242424") 
        report_frame.place(x=470, y=420)
        
        #Thêm nút xuất báo cáo tác giả và chức năng xuất ra file excel số tác phẩm theo tác giả
        #Sử dụng thêm thư viện để có thể cho người dùng mở cửa sổ Explorer để chọn nơi lưu file

        from tkinter import filedialog
        import openpyxl  # Đổi sang thư viện openpyxl
        def select_save_location():
            root = tk.Tk()
            root.withdraw()  # Ẩn cửa sổ gốc
            file_path = filedialog.asksaveasfilename(
                title="Select file",
                defaultextension=".xlsx",
                filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
            )
            return file_path

        # Function to generate report based on selected report type
        def generate_report():
            report_type = report_type_combobox.get()
            file_path = select_save_location()
            
            if file_path:
                # Connect to the MySQL database
                mydb = mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="130603",
                    database="QLThuVien"
                )
                mycursor = mydb.cursor()
                
                # Choose query based on selected report type
                if report_type == "Báo cáo Sách theo Tác Giả":
                    query = """
                        SELECT Sach.MaSach, Sach.TenSach, TacGia.TenTG
                        FROM Sach
                        JOIN TacGia ON Sach.MaTG = TacGia.MaTG
                        ORDER BY TacGia.TenTG
                    """
                    headers = ["Mã Sách", "Tên Sách", "Tác Giả"]
                
                elif report_type == "Báo cáo Sách theo Nhà Xuất Bản":
                    query = """
                        SELECT Sach.MaSach, Sach.TenSach, NhaXuatBan.TenNXB
                        FROM Sach
                        JOIN NhaXuatBan ON Sach.MaNXB = NhaXuatBan.MaNXB
                        ORDER BY NhaXuatBan.TenNXB
                    """
                    headers = ["Mã Sách", "Tên Sách", "Nhà Xuất Bản"]
                
                elif report_type == "Báo cáo Sách theo Thể Loại":
                    query = """
                        SELECT Sach.MaSach, Sach.TenSach, TheLoai.TenTL
                        FROM Sach
                        JOIN TheLoai ON Sach.MaTL = TheLoai.MaTL
                        ORDER BY TheLoai.TenTL
                    """
                    headers = ["Mã Sách", "Tên Sách", "Thể Loại"]
                
                elif report_type == "Danh Sách Tất Cả Sách":
                    query = """
                        SELECT MaSach, TenSach, MaTG, MaTL, MaNXB, NamXB, SoLuong
                        FROM Sach
                    """
                    headers = ["Mã Sách", "Tên Sách", "Mã Tác Giả", "Mã Thể Loại", "Mã NXB", "Năm XB", "Số Lượng"]
                
                else:
                    tk.messagebox.showwarning("Warning", "Vui lòng chọn thể loại báo cáo!")
                    return
                
                # Execute query and fetch data
                mycursor.execute(query)
                myresult = mycursor.fetchall()

                # Create Excel workbook and write data
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = report_type

                # Write headers to Excel sheet
                worksheet.append(headers)

                # Write rows to Excel sheet
                for row in myresult:
                    worksheet.append(row)

                # Save Excel file
                workbook.save(file_path)
                
                # Close database connection
                mydb.close()
                tk.messagebox.showinfo("Success", f"{report_type} đã được xuất thành công!")

        # Add ComboBox for report type selection
        report_types = ["Báo cáo Sách theo Tác Giả", "Báo cáo Sách theo Nhà Xuất Bản", "Báo cáo Sách theo Thể Loại", "Danh Sách Tất Cả Sách"]
        report_type_combobox = ctk.CTkComboBox(
            report_frame, values=report_types, state="readonly",
            bg_color="#242424", border_color="#bcd5e8", corner_radius=10,
            button_color="#bcd5e8", button_hover_color="#60797d",
            dropdown_fg_color="#bcd5e8", dropdown_hover_color="#60797d",
            width=200
        )
        report_type_combobox.set("Chọn loại báo cáo")
        report_type_combobox.place(x=50, y=20)

        # Add button to generate report
        report_button = ctk.CTkButton(report_frame, text="Xuất báo cáo", width=150, command=generate_report)
        report_button.place(x=75, y=80)
   
    def switch_to_publisher_management():
        def execute_query(query, params=None):
            try:
                with mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="130603",
                    database="QLThuVien"
                ) as mydb:
                    cursor = mydb.cursor()
                    cursor.execute(query, params or ())
                    return cursor.fetchall()
            except mysql.connector.Error as err:
                messagebox.showerror("Database Error", f"Lỗi cơ sở dữ liệu: {err}")
                return None
            
        sort_direction = {"AZ": "ASC", "ZA": "DESC"}
        
        def search_publisher(search_entry, combobox):
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if combobox.get() == "Mã NXB":
                mycursor.execute("SELECT * FROM NhaXuatBan WHERE MaNXB = %s", (search_entry.get(),))
            elif combobox.get() == "Tên NXB":
                mycursor.execute("SELECT * FROM NhaXuatBan WHERE TenNXB = %s", (search_entry.get(),))
            elif combobox.get() == "Địa Chỉ":
                mycursor.execute("SELECT * FROM NhaXuatBan WHERE DiaChi = %s", (search_entry.get(),))
            elif combobox.get() == "Email":
                mycursor.execute("SELECT * FROM NhaXuatBan WHERE Email = %s", (search_entry.get(),))
            elif combobox.get() == "Người Đại Diện":
                mycursor.execute("SELECT * FROM NhaXuatBan WHERE NguoiDaiDien = %s", (search_entry.get(),))
            myresult = mycursor.fetchall()
            for i in table.get_children():
                table.delete(i)
            for row in myresult:
                table.insert("", "end", values=row)
            mydb.close()

        def add_publisher():
            pass

        label = ctk.CTkLabel(content_frame, text="Quản lí nhà xuất bản", font=("Century Gothic", 20), text_color="white")
        label.pack(pady=20)
        label.place(x=340, y=10)

        nxb = ["Mã NXB", "Tên NXB", "Địa Chỉ", "Email", "Người Đại Diện"]
        combobox = ctk.CTkComboBox(content_frame, values=nxb, bg_color="#242424", border_color="#bcd5e8", corner_radius=10, button_color="#bcd5e8", button_hover_color="#60797d", dropdown_fg_color="#bcd5e8", dropdown_hover_color="#60797d", width=150, state="readonly")
        combobox.place(x=100, y=60)
        combobox.set("Tìm kiếm bằng")
        search_entry = ctk.CTkEntry(content_frame, width=300, placeholder_text="")
        search_entry.pack(pady=20)
        search_entry.place(x=260, y=60)

        search_button = ctk.CTkButton(content_frame, text="Tìm kiếm", width=50, command=lambda: search_publisher(search_entry, combobox))
        search_button.pack(pady=20)
        search_button.place(x=570, y=60)

        def show_all_publisher():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            mycursor.execute("SELECT * FROM NhaXuatBan")
            myresult = mycursor.fetchall()
            for i in table.get_children():
                table.delete(i)
            for row in myresult:
                table.insert("", "end", values=row)
            mydb.close()

        show_all_button = ctk.CTkButton(content_frame, text="Hiển thị tất cả", width=50, command=show_all_publisher)
        show_all_button.pack(pady=20)
        show_all_button.place(x=650, y=60)

        columns = ("Mã NXB", "Tên NXB", "Địa Chỉ", "Email", "Người Đại Diện")
        table = ttk.Treeview(content_frame, columns=columns, show="headings", height=11)
        table.place(x=117, y=120)
        for col in columns:
            table.heading(col, text=col)
        table.column("Mã NXB", width=100)
        table.column("Người Đại Diện", width=300)
        for col in columns:
            table.column(col, anchor="center")

        scroll = ttk.Scrollbar(content_frame, orient="vertical", command=table.yview)
        scroll.place(x=920, y=120, height=247)
        table.configure(yscrollcommand=scroll.set)
        show_all_publisher()

        info_frame = ctk.CTkFrame(content_frame, width=350, height=300, corner_radius=15, bg_color="#242424", fg_color="white")
        info_frame.place(x=95, y=300)

        label = ctk.CTkLabel(info_frame, text="Thông tin nhà xuất bản", font=("Century Gothic", 20), text_color="black")
        label.pack(pady=20)
        label.place(x=95, y=10)

        label = ctk.CTkLabel(info_frame, text="Mã NXB", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=40)
        maNXB_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        maNXB_entry.pack(pady=10)
        maNXB_entry.place(x=10, y=70)

        label = ctk.CTkLabel(info_frame, text="Tên NXB", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=100)
        tenNXB_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        tenNXB_entry.pack(pady=10)
        tenNXB_entry.place(x=10, y=130)

        label = ctk.CTkLabel(info_frame, text="Địa Chỉ", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=160)
        diaChi_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        diaChi_entry.pack(pady=10)
        diaChi_entry.place(x=10, y=190)

        label = ctk.CTkLabel(info_frame, text="Email", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=220)
        email_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        email_entry.pack(pady=10)
        email_entry.place(x=10, y=250)

        label = ctk.CTkLabel(info_frame, text="Người Đại Diện", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=280)
        nguoiDaiDien_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        nguoiDaiDien_entry.pack(pady=10)
        nguoiDaiDien_entry.place(x=10, y=310)

        def on_click(event):
            item = table.selection()[0]
            maNXB_entry.delete(0, "end")
            tenNXB_entry.delete(0, "end")
            diaChi_entry.delete(0, "end")
            email_entry.delete(0, "end")
            nguoiDaiDien_entry.delete(0, "end")
            maNXB_entry.insert(0, table.item(item, "values")[0])
            tenNXB_entry.insert(0, table.item(item, "values")[1])
            diaChi_entry.insert(0, table.item(item, "values")[2])
            email_entry.insert(0, table.item(item, "values")[3])
            nguoiDaiDien_entry.insert(0, table.item(item, "values")[4])

        table.bind("<ButtonRelease-1>", on_click)

        button_frame = ctk.CTkFrame(content_frame, width=300, height=120, corner_radius=15, bg_color="#242424", fg_color="#242424")
        button_frame.place(x=500, y=300)

        def add_publisher():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if not maNXB_entry.get() or not tenNXB_entry.get() or not diaChi_entry.get() or not email_entry.get() or not nguoiDaiDien_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập đủ thông tin!")
            else:
                try:
                    mycursor.execute("INSERT INTO NhaXuatBan (MaNXB, TenNXB, DiaChi, Email, NguoiDaiDien) VALUES (%s, %s, %s, %s, %s)", (maNXB_entry.get(), tenNXB_entry.get(), diaChi_entry.get(), email_entry.get(), nguoiDaiDien_entry.get()))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Thêm nhà xuất bản thành công!")
                    show_all_publisher()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã nhà xuất bản đã tồn tại!")
            mydb.close()
        add_button = ctk.CTkButton(button_frame, text="Thêm", width=100, command=add_publisher)
        add_button.pack(pady=30)
        add_button.place(x=0, y=25)

        def update_publisher():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if not maNXB_entry.get() or not tenNXB_entry.get() or not diaChi_entry.get() or not email_entry.get() or not nguoiDaiDien_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập đầy đủ thông tin nhà xuất bản cần cập nhật!")
            else:
                try:
                    mycursor.execute("UPDATE NhaXuatBan SET TenNXB = %s, DiaChi = %s, Email = %s, NguoiDaiDien = %s WHERE MaNXB = %s", (tenNXB_entry.get(), diaChi_entry.get(), email_entry.get(), nguoiDaiDien_entry.get(), maNXB_entry.get()))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Cập nhật thông tin nhà xuất bản thành công!")
                    show_all_publisher()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã nhà xuất bản không tồn tại!")
            mydb.close()
        update_button = ctk.CTkButton(button_frame, text="Cập nhật", width=100, command=update_publisher)
        update_button.pack(pady=30)
        update_button.place(x=140, y=25)

        def delete_publisher():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if not maNXB_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập mã nhà xuất bản cần xóa!")
            else:
                try:
                    mycursor.execute("DELETE FROM NhaXuatBan WHERE MaNXB = %s", (maNXB_entry.get(),))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Xóa nhà xuất bản thành công!")
                    show_all_publisher()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã nhà xuất bản không tồn tại!")
            mydb.close()
            reset_info()
        delete_button = ctk.CTkButton(button_frame, text="Xóa", width=100, command=delete_publisher)
        delete_button.pack(pady=30)
        delete_button.place(x=140, y=80)

        def reset_info():
            maNXB_entry.delete(0, "end")
            tenNXB_entry.delete(0, "end")
            diaChi_entry.delete(0, "end")
            email_entry.delete(0, "end")
            nguoiDaiDien_entry.delete(0, "end")
        reset_button = ctk.CTkButton(button_frame, text="Reset", width=100, command=reset_info)
        reset_button.pack(pady=30)
        reset_button.place(x=0, y=80)

        report_frame = ctk.CTkFrame(content_frame, width=300, height=150, corner_radius=15, bg_color="#242424", fg_color="white")
        report_frame.place(x=470, y=420)

        from tkinter import filedialog
        import openpyxl
        def select_save_location():
            root = tk.Tk()
            root.withdraw()
            file_path = filedialog.asksaveasfilename(
                title="Select file",
                defaultextension=".xlsx",
                filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
            )
            return file_path

        def report_publisher():
            file_path = select_save_location()
            if file_path:
                mydb = mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="130603",
                    database="QLThuVien"
                )
                mycursor = mydb.cursor()
                mycursor.execute("""
                    SELECT NhaXuatBan.MaNXB, NhaXuatBan.TenNXB, COUNT(Sach.MaNXB) AS SoLuongSach
                    FROM NhaXuatBan 
                    JOIN Sach ON NhaXuatBan.MaNXB = Sach.MaNXB 
                    GROUP BY NhaXuatBan.MaNXB
                """)
                myresult = mycursor.fetchall()

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "Báo cáo Nhà Xuất Bản"

                headers = ["Mã NXB", "Tên NXB", "Số Tác Phẩm"]
                worksheet.append(headers)

                for result in myresult:
                    worksheet.append(result)

                workbook.save(file_path)
                mydb.close()

        report_button = ctk.CTkButton(report_frame, text="Xuất báo cáo nhà xuất bản", width=150, command=report_publisher)
        report_button.place(x=75, y=0)

    def switch_to_book_management():
        # Function to execute query and return results
        def search_book(search_entry, combobox, table):
            try:
                # Establish a database connection
                mydb = mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="130603",
                    database="qltv"
                )
                mycursor = mydb.cursor()

                # Prepare search value with wildcards
                search_value = '%' + search_entry.get() + '%'

                # Define the search query based on combobox selection
                query = ""
                if combobox.get() == "Mã Sách":
                    query = "SELECT * FROM Sach WHERE MaSach LIKE %s"
                elif combobox.get() == "Tên Sách":
                    query = "SELECT * FROM Sach WHERE TenSach LIKE %s"
                elif combobox.get() == "Mã Tác Giả":
                    query = "SELECT * FROM Sach WHERE MaTG LIKE %s"
                elif combobox.get() == "Mã Thể Loại":
                    query = "SELECT * FROM Sach WHERE MaTL LIKE %s"
                elif combobox.get() == "Mã NXB":
                    query = "SELECT * FROM Sach WHERE MaNXB LIKE %s"
                elif combobox.get() == "Năm XB":
                    query = "SELECT * FROM Sach WHERE NamXB LIKE %s"

                # Execute the query with the search value
                mycursor.execute(query, (search_value,))
                results = mycursor.fetchall()  # Fetch all results to clear the cursor

                # Clear the table before inserting new data
                for row in table.get_children():
                    table.delete(row)

                # Insert search results into the table
                for result in results:
                    table.insert("", "end", values=result)

            except mysql.connector.Error as err:
                messagebox.showerror("Database Error", f"Lỗi cơ sở dữ liệu: {err}")

            finally:
                # Close cursor and database connection if open
                if 'mycursor' in locals():
                    mycursor.close()
                if 'mydb' in locals() and mydb.is_connected():
                    mydb.close()
                    
        # show all books in the database
        def show_all_books():
            try:
                mydb = mysql.connector.connect(
                    host="localhost",
                    username="root",
                    password="130603",
                    database="qltv"
                )
                mycursor = mydb.cursor()
                mycursor.execute("SELECT * FROM Sach")
                results = mycursor.fetchall()
                for row in table.get_children():
                    table.delete(row)
                for result in results:
                    table.insert("", "end", values=result)
            except mysql.connector.Error as err:
                messagebox.showerror("Database Error", f"Lỗi cơ sở dữ liệu: {err}")
            finally:
                if mycursor:
                    mycursor.close()
                if mydb:
                    mydb.close()
                    
        label = ctk.CTkLabel(content_frame, text="Quản lí sách", font=("Century Gothic", 20), text_color="white")
        label.pack(pady=20)
        label.place(x=340, y=10)

        sach = ["Mã Sách", "Tên Sách", "Mã Tác Giả", "Mã Thể Loại", "Mã NXB", "Năm XB"]
        combobox = ctk.CTkComboBox(content_frame, values=sach, bg_color="#242424", border_color="#bcd5e8", corner_radius=10, button_color="#bcd5e8", button_hover_color="#60797d", dropdown_fg_color="#bcd5e8", dropdown_hover_color="#60797d", width=150, state="readonly")
        combobox.place(x=100, y=60)
        combobox.set("Tìm kiếm bằng")
        search_entry = ctk.CTkEntry(content_frame, width=300, placeholder_text="")
        search_entry.pack(pady=20)
        search_entry.place(x=260, y=60)

        search_button = ctk.CTkButton(content_frame, text="Tìm kiếm", width=50, command=lambda: search_book(search_entry, combobox, table))
        search_button.pack(pady=20)
        search_button.place(x=570, y=60)

                
        show_all_button = ctk.CTkButton(content_frame, text="Hiển thị tất cả", width=50, command=show_all_books)
        show_all_button.pack(pady=20)
        show_all_button.place(x=650, y=60)

        # Set up the table frame
        table_frame = ctk.CTkFrame(content_frame, width=700, fg_color = "#242424")
        table_frame.place(x=60, y=100)

        # Define columns for the table
        columns = ("Mã Sách", "Tên Sách", "Mã Tác Giả", "Mã Thể Loại", "Mã NXB", "Năm XB", "Số Lượng", "Số lượng còn lại")
        table = ttk.Treeview(table_frame, columns=columns, show="headings", height=11)
        table.place(x=0, y=0)

        # Set up column headings and widths
        for col in columns:
            table.heading(col, text=col)
            table.column(col, anchor="center")

        # Customize widths for each column based on content
        table.column("Mã Sách", width=80)            # Reduced width for compact ID display
        table.column("Tên Sách", width=250)           # Increased width for longer book names
        table.column("Mã Tác Giả", width=90)          # Reduced slightly
        table.column("Mã Thể Loại", width=90)         # Reduced slightly
        table.column("Mã NXB", width=90)              # Reduced slightly
        table.column("Năm XB", width=80)              # Smaller width for year
        table.column("Số Lượng", width=90)            # Standard width for quantity
        table.column("Số lượng còn lại", width=110)   # Slightly wider for better visibility

        scroll = ttk.Scrollbar(content_frame, orient="vertical", command=table.yview)
        scroll.place(x=955, y=125, height=247)
        table.configure(yscrollcommand=scroll.set)
        
        info_frame = ctk.CTkFrame(content_frame, width=350, height=600, corner_radius=15, bg_color="#242424", fg_color="#242424")
        info_frame.place(x=95, y=300)

        show_all_books()
        label = ctk.CTkLabel(info_frame, text="Thông tin sách", font=("Century Gothic", 20), text_color="white")
        label.pack(pady=20)
        label.place(x=95, y=10)

        label = ctk.CTkLabel(info_frame, text="Mã Sách", font=("Century Gothic", 15), text_color="white")
        label.pack(pady=10)
        label.place(x=10, y=40)
        maSach_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        maSach_entry.pack(pady=10)
        maSach_entry.place(x=10, y=70)

        label = ctk.CTkLabel(info_frame, text="Tên Sách", font=("Century Gothic", 15), text_color="white")
        label.pack(pady=10)
        label.place(x=10, y=100)
        tenSach_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        tenSach_entry.pack(pady=10)
        tenSach_entry.place(x=10, y=130)

        label = ctk.CTkLabel(info_frame, text="Mã Tác Giả", font=("Century Gothic", 15), text_color="white")
        label.pack(pady=10)
        label.place(x=10, y=160)
        maTG_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        maTG_entry.pack(pady=10)
        maTG_entry.place(x=10, y=190)

        label = ctk.CTkLabel(info_frame, text="Mã Thể Loại", font=("Century Gothic", 15), text_color="white")
        label.pack(pady=10)
        label.place(x=10, y=220)
        maTL_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        maTL_entry.pack(pady=10)
        maTL_entry.place(x=10, y=250)

        label = ctk.CTkLabel(info_frame, text="Mã NXB", font=("Century Gothic", 15), text_color="white")
        label.pack(pady=10)
        label.place(x=10, y=280)
        maNXB_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        maNXB_entry.pack(pady=10)
        maNXB_entry.place(x=10, y=310)

        label = ctk.CTkLabel(info_frame, text="Năm XB", font=("Century Gothic", 15), text_color="white")
        label.pack(pady=10)
        label.place(x=10, y=340)
        namXB_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        namXB_entry.pack(pady=10)
        namXB_entry.place(x=10, y=370)
        
        label = ctk.CTkLabel(info_frame, text="Số Lượng", font=("Century Gothic", 15), text_color="white")
        label.pack(pady=10)
        label.place(x=10, y=400)
        soLuong_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        soLuong_entry.pack(pady=10)
        soLuong_entry.place(x=10, y=430)

        def on_click(event):
            item = table.selection()[0]
            maSach_entry.delete(0, "end")
            tenSach_entry.delete(0, "end")
            maTG_entry.delete(0, "end")
            maTL_entry.delete(0, "end")
            maNXB_entry.delete(0, "end")
            namXB_entry.delete(0, "end")
            soLuong_entry.delete(0, "end")
            maSach_entry.insert(0, table.item(item, "values")[0])
            tenSach_entry.insert(0, table.item(item, "values")[1])
            maTG_entry.insert(0, table.item(item, "values")[2])
            maTL_entry.insert(0, table.item(item, "values")[3])
            maNXB_entry.insert(0, table.item(item, "values")[4])
            namXB_entry.insert(0, table.item(item, "values")[5])
            soLuong_entry.insert(0, table.item(item, "values")[6])

        table.bind("<ButtonRelease-1>", on_click)

        button_frame = ctk.CTkFrame(content_frame, width=300, height=120, corner_radius=15, bg_color="#242424", fg_color="#242424")
        button_frame.place(x=500, y=350)

        def add_book():
            try:
                mydb = mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="130603",
                    database="qltv"
                )
                mycursor = mydb.cursor()
                mycursor.callproc('ThemSach', [
                    maSach_entry.get(),
                    tenSach_entry.get(),
                    maTG_entry.get(),
                    maTL_entry.get(),
                    maNXB_entry.get(),
                    namXB_entry.get(),
                    soLuong_entry.get()
                ])
                mydb.commit()
                messagebox.showinfo("Success", "Thêm sách thành công!")
                show_all_books()
            except mysql.connector.Error as err:
                if err.errno == 1644:
                    messagebox.showerror("Database Error", "Mã sách đã tồn tại!")
                else:
                    messagebox.showerror("Database Error", f"Lỗi cơ sở dữ liệu: {err}")
            finally:
                if mycursor:
                    mycursor.close()
                if mydb:
                    mydb.close()
        add_button = ctk.CTkButton(button_frame, text="Thêm", width=100, command=add_book)
        add_button.pack(pady=30)
        add_button.place(x=0, y=25)

        def update_book():
            try:
                # Chuyển đổi năm xuất bản và tổng số lượng thành số nguyên
                nam_xb = int(namXB_entry.get())
                tong_so_luong = int(soLuong_entry.get())
                
                # Kết nối đến cơ sở dữ liệu
                mydb = mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="130603",
                    database="qltv"
                )
                mycursor = mydb.cursor()

                # Gọi stored procedure SuaSach với các tham số cập nhật
                mycursor.callproc('SuaSach', [
                    maSach_entry.get(),
                    tenSach_entry.get(),
                    maTG_entry.get(),
                    maTL_entry.get(),
                    maNXB_entry.get(),
                    nam_xb,         # Chuyển đổi thành số nguyên
                    tong_so_luong   # Chuyển đổi thành số nguyên
                ])
                mydb.commit()

                # Thông báo cập nhật thành công
                messagebox.showinfo("Success", "Cập nhật sách thành công!")
                show_all_books()
            
            except ValueError:
                # Xử lý lỗi nếu chuyển đổi giá trị không thành công
                messagebox.showerror("Input Error", "Vui lòng nhập số hợp lệ cho Năm XB và Tổng Số Lượng.")
            
            except mysql.connector.Error as err:
                # Kiểm tra thông báo lỗi thay vì errno
                error_message = str(err)
                if "Sách không tồn tại trong hệ thống" in error_message:
                    messagebox.showerror("Database Error", "Mã sách không tồn tại!")
                elif "Tổng số lượng sách không thể nhỏ hơn số lượng sách đang cho mượn" in error_message:
                    messagebox.showerror("Database Error", "Tổng số lượng sách không thể nhỏ hơn số lượng sách đang cho mượn hoặc đã được xác nhận!")
                else:
                    messagebox.showerror("Database Error", f"Lỗi cơ sở dữ liệu: {err}")

            finally:
                # Đảm bảo đóng kết nối và con trỏ nếu được tạo
                if 'mycursor' in locals() and mycursor:
                    mycursor.close()
                if 'mydb' in locals() and mydb.is_connected():
                    mydb.close()
                    
        update_button = ctk.CTkButton(button_frame, text="Cập nhật", width=100, command=update_book)
        update_button.pack(pady=30)
        update_button.place(x=140, y=25)

        def delete_book():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="qltv"
            )
            mycursor = mydb.cursor()
            if not maSach_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập mã sách cần xóa!")
            else:
                try:
                    mycursor.callproc('XoaSach', [maSach_entry.get()])
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Xóa sách thành công!")
                    show_all_books()
                except mysql.connector.Error as err:
                    if err.errno == 1644:
                        if "Sách đang được mượn" in str(err):
                            tk.messagebox.showwarning("Warning", "Sách đang được mượn! Bạn không thể xóa sách này!")
                        else:
                            tk.messagebox.showwarning("Warning", "Mã sách không tồn tại!")
                finally:
                    if mycursor:
                        mycursor.close()
                    if mydb:
                        mydb.close()
                        
        delete_button = ctk.CTkButton(button_frame, text="Xóa", width=100, command=delete_book)
        delete_button.pack(pady=30)
        delete_button.place(x=140, y=80)

        def reset_info():
            maSach_entry.delete(0, "end")
            tenSach_entry.delete(0, "end")
            maTG_entry.delete(0, "end")
            maTL_entry.delete(0, "end")
            maNXB_entry.delete(0, "end")
            namXB_entry.delete(0, "end")
        reset_button = ctk.CTkButton(button_frame, text="Reset", width=100, command=reset_info)
        reset_button.pack(pady=30)
        reset_button.place(x=0, y=80)

        report_frame = ctk.CTkFrame(content_frame, width=300, height=150, corner_radius=15, bg_color="#242424", fg_color="#00ac47")
        report_frame.place(x=470, y=500)

        from tkinter import filedialog
        import openpyxl
        def select_save_location():
            root = tk.Tk()
            root.withdraw()
            file_path = filedialog.asksaveasfilename(
                title="Select file",
                defaultextension=".xlsx",
                filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
            )
            return file_path

        def report_book():
            file_path = select_save_location()
            if file_path:
                mydb = mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="130603",
                    database="QLThuVien"
                )
                mycursor = mydb.cursor()
                mycursor.execute("""
                    SELECT Sach.MaSach, Sach.TenSach, TacGia.TenTG, TheLoai.TenTL, NhaXuatBan.TenNXB, Sach.NamXB
                    FROM Sach 
                    JOIN TacGia ON Sach.MaTG = TacGia.MaTG
                    JOIN TheLoai ON Sach.MaTL = TheLoai.MaTL
                    JOIN NhaXuatBan ON Sach.MaNXB = NhaXuatBan.MaNXB
                """)
                myresult = mycursor.fetchall()

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "Báo cáo Sách"

                headers = ["Mã Sách", "Tên Sách", "Tên Tác Giả", "Tên Thể Loại", "Tên NXB", "Năm XB"]
                worksheet.append(headers)

                for result in myresult:
                    worksheet.append(result)

                workbook.save(file_path)
                mydb.close()

        report_button = ctk.CTkButton(report_frame, text="Xuất báo cáo sách", width=150, command=report_book)
        report_button.place(x=75, y=20)

    def switch_to_borrow_return_management():
        label = ctk.CTkLabel(content_frame, text="Quản lí mượn trả sách", font=("Century Gothic", 20))
        label.pack(pady=20)
        # Add more widgets for borrow/return management

    def switch_to_staff_management():
        def search_staff(search_entry, combobox):
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if combobox.get() == "Mã NV":
                mycursor.execute("SELECT * FROM NhanVien WHERE MaNV = %s", (search_entry.get(),))
            elif combobox.get() == "Tên NV":
                mycursor.execute("SELECT * FROM NhanVien WHERE TenNV = %s", (search_entry.get(),))
            elif combobox.get() == "Ngày Sinh":
                mycursor.execute("SELECT * FROM NhanVien WHERE NgaySinh = %s", (search_entry.get(),))
            elif combobox.get() == "SĐT":
                mycursor.execute("SELECT * FROM NhanVien WHERE SDT = %s", (search_entry.get(),))
            myresult = mycursor.fetchall()
            for i in table.get_children():
                table.delete(i)
            for row in myresult:
                table.insert("", "end", values=row)
            mydb.close()

        def add_staff():
            pass

        label = ctk.CTkLabel(content_frame, text="Quản lí nhân viên", font=("Century Gothic", 20), text_color="white")
        label.pack(pady=20)
        label.place(x=340, y=10)

        nv = ["Mã NV", "Tên NV", "Ngày Sinh", "SĐT"]
        combobox = ctk.CTkComboBox(content_frame, values=nv, bg_color="#242424", border_color="#bcd5e8", corner_radius=10, button_color="#bcd5e8", button_hover_color="#60797d", dropdown_fg_color="#bcd5e8", dropdown_hover_color="#60797d", width=150, state="readonly")
        combobox.place(x=100, y=60)
        combobox.set("Tìm kiếm bằng")
        search_entry = ctk.CTkEntry(content_frame, width=300, placeholder_text="")
        search_entry.pack(pady=20)
        search_entry.place(x=260, y=60)

        search_button = ctk.CTkButton(content_frame, text="Tìm kiếm", width=50, command=lambda: search_staff(search_entry, combobox))
        search_button.pack(pady=20)
        search_button.place(x=570, y=60)

        def show_all_staff():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            mycursor.execute("SELECT * FROM NhanVien")
            myresult = mycursor.fetchall()
            for i in table.get_children():
                table.delete(i)
            for row in myresult:
                table.insert("", "end", values=row)
            mydb.close()

        show_all_button = ctk.CTkButton(content_frame, text="Hiển thị tất cả", width=50, command=show_all_staff)
        show_all_button.pack(pady=20)
        show_all_button.place(x=650, y=60)

        columns = ("Mã NV", "Tên NV", "Ngày Sinh", "SĐT")
        table = ttk.Treeview(content_frame, columns=columns, show="headings", height=11)
        table.place(x=117, y=120)
        for col in columns:
            table.heading(col, text=col)
        table.column("Mã NV", width=100)
        table.column("SĐT", width=150)
        for col in columns:
            table.column(col, anchor="center")

        scroll = ttk.Scrollbar(content_frame, orient="vertical", command=table.yview)
        scroll.place(x=920, y=120, height=247)
        table.configure(yscrollcommand=scroll.set)
        show_all_staff()

        info_frame = ctk.CTkFrame(content_frame, width=350, height=300, corner_radius=15, bg_color="#242424", fg_color="white")
        info_frame.place(x=95, y=300)

        label = ctk.CTkLabel(info_frame, text="Thông tin nhân viên", font=("Century Gothic", 20), text_color="black")
        label.pack(pady=20)
        label.place(x=95, y=10)

        label = ctk.CTkLabel(info_frame, text="Mã NV", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=40)
        maNV_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        maNV_entry.pack(pady=10)
        maNV_entry.place(x=10, y=70)

        label = ctk.CTkLabel(info_frame, text="Tên NV", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=100)
        tenNV_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        tenNV_entry.pack(pady=10)
        tenNV_entry.place(x=10, y=130)

        label = ctk.CTkLabel(info_frame, text="Ngày Sinh", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=160)
        ngaySinh_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        ngaySinh_entry.pack(pady=10)
        ngaySinh_entry.place(x=10, y=190)

        label = ctk.CTkLabel(info_frame, text="SĐT", font=("Century Gothic", 15), text_color="black")
        label.pack(pady=10)
        label.place(x=10, y=220)
        sdt_entry = ctk.CTkEntry(info_frame, width=320, placeholder_text="")
        sdt_entry.pack(pady=10)
        sdt_entry.place(x=10, y=250)

        def on_click(event):
            item = table.selection()[0]
            maNV_entry.delete(0, "end")
            tenNV_entry.delete(0, "end")
            ngaySinh_entry.delete(0, "end")
            sdt_entry.delete(0, "end")
            maNV_entry.insert(0, table.item(item, "values")[0])
            tenNV_entry.insert(0, table.item(item, "values")[1])
            ngaySinh_entry.insert(0, table.item(item, "values")[2])
            sdt_entry.insert(0, table.item(item, "values")[3])

        table.bind("<ButtonRelease-1>", on_click)

        button_frame = ctk.CTkFrame(content_frame, width=300, height=120, corner_radius=15, bg_color="#242424", fg_color="#242424")
        button_frame.place(x=500, y=300)

        def add_staff():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if not maNV_entry.get() or not tenNV_entry.get() or not ngaySinh_entry.get() or not sdt_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập đủ thông tin!")
            else:
                try:
                    mycursor.execute("INSERT INTO NhanVien (MaNV, TenNV, NgaySinh, SDT) VALUES (%s, %s, %s, %s)", (maNV_entry.get(), tenNV_entry.get(), ngaySinh_entry.get(), sdt_entry.get()))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Thêm nhân viên thành công!")
                    show_all_staff()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã nhân viên đã tồn tại!")
            mydb.close()
        add_button = ctk.CTkButton(button_frame, text="Thêm", width=100, command=add_staff)
        add_button.pack(pady=30)
        add_button.place(x=0, y=25)

        def update_staff():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if not maNV_entry.get() or not tenNV_entry.get() or not ngaySinh_entry.get() or not sdt_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập đầy đủ thông tin nhân viên cần cập nhật!")
            else:
                try:
                    mycursor.execute("UPDATE NhanVien SET TenNV = %s, NgaySinh = %s, SDT = %s WHERE MaNV = %s", (tenNV_entry.get(), ngaySinh_entry.get(), sdt_entry.get(), maNV_entry.get()))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Cập nhật thông tin nhân viên thành công!")
                    show_all_staff()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã nhân viên không tồn tại!")
            mydb.close()
        update_button = ctk.CTkButton(button_frame, text="Cập nhật", width=100, command=update_staff)
        update_button.pack(pady=30)
        update_button.place(x=140, y=25)

        def delete_staff():
            mydb = mysql.connector.connect(
                host="localhost",
                user="root",
                password="130603",
                database="QLThuVien"
            )
            mycursor = mydb.cursor()
            if not maNV_entry.get():
                tk.messagebox.showwarning("Warning", "Vui lòng nhập mã nhân viên cần xóa!")
            else:
                try:
                    mycursor.execute("DELETE FROM NhanVien WHERE MaNV = %s", (maNV_entry.get(),))
                    mydb.commit()
                    tk.messagebox.showinfo("Success", "Xóa nhân viên thành công!")
                    show_all_staff()
                except mysql.connector.IntegrityError:
                    tk.messagebox.showwarning("Warning", "Mã nhân viên không tồn tại!")
            mydb.close()
            reset_info()
        delete_button = ctk.CTkButton(button_frame, text="Xóa", width=100, command=delete_staff)
        delete_button.pack(pady=30)
        delete_button.place(x=140, y=80)

        def reset_info():
            maNV_entry.delete(0, "end")
            tenNV_entry.delete(0, "end")
            ngaySinh_entry.delete(0, "end")
            sdt_entry.delete(0, "end")
        reset_button = ctk.CTkButton(button_frame, text="Reset", width=100, command=reset_info)
        reset_button.pack(pady=30)
        reset_button.place(x=0, y=80)

        report_frame = ctk.CTkFrame(content_frame, width=300, height=150, corner_radius=15, bg_color="#242424", fg_color="white")
        report_frame.place(x=470, y=420)

        from tkinter import filedialog
        import openpyxl
        def select_save_location():
            root = tk.Tk()
            root.withdraw()
            file_path = filedialog.asksaveasfilename(
                title="Select file",
                defaultextension=".xlsx",
                filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
            )
            return file_path

        def report_staff():
            file_path = select_save_location()
            if file_path:
                mydb = mysql.connector.connect(
                    host="localhost",
                    user="root",
                    password="130603",
                    database="QLThuVien"
                )
                mycursor = mydb.cursor()
                mycursor.execute("""
                    SELECT MaNV, TenNV, NgaySinh, SDT
                    FROM NhanVien
                """)
                myresult = mycursor.fetchall()

                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = "Báo cáo Nhân Viên"

                headers = ["Mã NV", "Tên NV", "Ngày Sinh", "SĐT"]
                worksheet.append(headers)

                for result in myresult:
                    worksheet.append(result)

                workbook.save(file_path)
                mydb.close()

        report_button = ctk.CTkButton(report_frame, text="Xuất báo cáo nhân viên", width=150, command=report_staff)
        report_button.place(x=75, y=0)

    # Function to handle combobox selection
    def on_combobox_select(choice):
        for widget in content_frame.winfo_children():
            widget.destroy()
        if choice == "Quản lí tác giả":
            switch_to_author_management()
        elif choice == "Quản lí nhà sản xuất":
            switch_to_publisher_management()
        elif choice == "Quản lí sách":
            switch_to_book_management()
        elif choice == "Quản lí mượn trả sách":
            switch_to_borrow_return_management()
        elif choice == "Quản lí nhân viên":
            switch_to_staff_management()
        elif choice == "Quản lí tài khoản":
            switch_to_account_management()
        elif choice == "Quản lí thông tin cá nhân":
            switch_to_personal_info()

    # Tạo combobox
    if role == "Admin":
        options = ["Quản lí nhân viên", "Quản lí tài khoản"]
    else:
        options = ["Quản lí sách","Quản lí tác giả", "Quản lí nhà sản xuất", "Quản lí thể loại",  "Quản lí mượn trả sách", "Quản lí thẻ thư viện", "Quản lí đọc giả", "Quản lí thông tin cá nhân"]
    combobox = ctk.CTkComboBox(
        app, values=options, 
        command=on_combobox_select, 
        bg_color="#2d4554", 
        border_color="#bcd5e8", 
        corner_radius=10,
        button_color="#bcd5e8",
        button_hover_color="#60797d",
        dropdown_fg_color="#bcd5e8",
        dropdown_hover_color="#60797d",
        width=170
    )
    combobox.place(x=0, y=97)
    combobox.set(options[0])
    on_combobox_select(options[0])  
    
    app.mainloop()

if __name__ == "__main__":
    main_app(username="admin", role = "NhanVien")  # Add username parameter